from django.contrib import admin
from simple_history.admin import SimpleHistoryAdmin
from import_export.admin import ImportExportModelAdmin
from .models import df_event_stts,df_fcst_ppltn,df_fcst12hours,df_fcst24hours,df_live_ppltn_stts,df_weather_stts

## 엑셀로 데이터 내보내기
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from import_export.admin import ExportActionMixin
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from datetime import datetime


## 행사정보
@admin.register(df_event_stts)
class MyModelAdmin_df_event_stts(ImportExportModelAdmin,SimpleHistoryAdmin):
    list_display = ['id','file_id', 'AREA_NM', 'EVENT_NM', 'EVENT_PERIOD','EVENT_PLACE'] ## 보여주기
    list_filter = ['AREA_NM'] ## 필터
    search_fields = ['AREA_NM', 'EVENT_NM', 'EVENT_PERIOD','EVENT_PLACE'] ## 검색
    list_per_page = 10 ##한페이지에 50개
    
    
    actions = ['export_selected']
    def export_selected(self, request, queryset):
        # queryset을 xlsx 파일로 내보냅니다.
        wb = Workbook()
        ws = wb.active

        # 필드명을 엑셀 첫 행에 추가합니다.
        field_names = [field.verbose_name.title() for field in queryset.model._meta.fields]
        ws.append(field_names)

        # 쿼리셋의 데이터를 엑셀 파일에 추가합니다.
        for obj in queryset:
            row_data = [getattr(obj, field.name) for field in queryset.model._meta.fields]
            ws.append(row_data)

        # 파일을 HttpResponse에 추가합니다.
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="selected_data_{datetime.today().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    export_selected.short_description = _("선택된 데이터 내보내기")  # 액션 버튼에 표시될 이름
    
## 시간별 인구 예상
@admin.register(df_fcst_ppltn)
class MyModelAdmin_df_fcst_ppltn(ImportExportModelAdmin,SimpleHistoryAdmin):
    list_display = ['id','file_id', 'AREA_NM', 'FCST_CONGEST_LVL', 'FCST_PPLTN_MIN','FCST_PPLTN_MAX'] ## 보여주기
    list_filter = ['AREA_NM'] ## 필터
    search_fields = ['AREA_NM'] ## 검색
    list_per_page = 10 ##한페이지에 50개
    
    
    actions = ['export_selected']
    def export_selected(self, request, queryset):
        # queryset을 xlsx 파일로 내보냅니다.
        wb = Workbook()
        ws = wb.active

        # 필드명을 엑셀 첫 행에 추가합니다.
        field_names = [field.verbose_name.title() for field in queryset.model._meta.fields]
        ws.append(field_names)

        # 쿼리셋의 데이터를 엑셀 파일에 추가합니다.
        for obj in queryset:
            row_data = [getattr(obj, field.name) for field in queryset.model._meta.fields]
            ws.append(row_data)

        # 파일을 HttpResponse에 추가합니다.
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="selected_data_{datetime.today().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    export_selected.short_description = _("선택된 데이터 내보내기")  # 액션 버튼에 표시될 이름
    
## 시간별 예상 날씨(12시간)
@admin.register(df_fcst12hours)
class MyModelAdmin_df_fcst12hours(ImportExportModelAdmin,SimpleHistoryAdmin):
    list_display = ['id','file_id', 'AREA_NM', 'FCST_DT', 'TEMP','PRECIPITATION','PRECPT_TYPE','RAIN_CHANCE','SKY_STTS'] ## 보여주기
    list_filter = ['AREA_NM'] ## 필터
    search_fields = ['AREA_NM'] ## 검색
    list_per_page = 10 ##한페이지에 50개
    
    
    actions = ['export_selected']
    def export_selected(self, request, queryset):
        # queryset을 xlsx 파일로 내보냅니다.
        wb = Workbook()
        ws = wb.active

        # 필드명을 엑셀 첫 행에 추가합니다.
        field_names = [field.verbose_name.title() for field in queryset.model._meta.fields]
        ws.append(field_names)

        # 쿼리셋의 데이터를 엑셀 파일에 추가합니다.
        for obj in queryset:
            row_data = [getattr(obj, field.name) for field in queryset.model._meta.fields]
            ws.append(row_data)

        # 파일을 HttpResponse에 추가합니다.
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="selected_data_{datetime.today().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    export_selected.short_description = _("선택된 데이터 내보내기")  # 액션 버튼에 표시될 이름
    
## 시간별 예상 날씨(24시간)
@admin.register(df_fcst24hours)
class MyModelAdmin_df_fcst24hours(ImportExportModelAdmin,SimpleHistoryAdmin):
    list_display = ['id','file_id', 'AREA_NM', 'FCST_DT', 'TEMP','PRECIPITATION','PRECPT_TYPE','RAIN_CHANCE','SKY_STTS'] ## 보여주기
    list_filter = ['AREA_NM'] ## 필터
    search_fields = ['AREA_NM'] ## 검색
    list_per_page = 10 ##한페이지에 50개
    
    
    actions = ['export_selected']
    def export_selected(self, request, queryset):
        # queryset을 xlsx 파일로 내보냅니다.
        wb = Workbook()
        ws = wb.active

        # 필드명을 엑셀 첫 행에 추가합니다.
        field_names = [field.verbose_name.title() for field in queryset.model._meta.fields]
        ws.append(field_names)

        # 쿼리셋의 데이터를 엑셀 파일에 추가합니다.
        for obj in queryset:
            row_data = [getattr(obj, field.name) for field in queryset.model._meta.fields]
            ws.append(row_data)

        # 파일을 HttpResponse에 추가합니다.
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="selected_data_{datetime.today().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    export_selected.short_description = _("선택된 데이터 내보내기")  # 액션 버튼에 표시될 이름
    
## 실시간인구현황
@admin.register(df_live_ppltn_stts)
class MyModelAdmin_df_live_ppltn_stts(ImportExportModelAdmin,SimpleHistoryAdmin):
    list_display = ['id','file_id', 'AREA_NM','AREA_CONGEST_MSG','AREA_PPLTN_MIN','AREA_PPLTN_MAX'] ## 보여주기
    list_filter = ['AREA_NM'] ## 필터
    search_fields = ['AREA_NM'] ## 검색
    list_per_page = 10 ##한페이지에 50개
    
    
    actions = ['export_selected']
    def export_selected(self, request, queryset):
        # queryset을 xlsx 파일로 내보냅니다.
        wb = Workbook()
        ws = wb.active

        # 필드명을 엑셀 첫 행에 추가합니다.
        field_names = [field.verbose_name.title() for field in queryset.model._meta.fields]
        ws.append(field_names)

        # 쿼리셋의 데이터를 엑셀 파일에 추가합니다.
        for obj in queryset:
            row_data = [getattr(obj, field.name) for field in queryset.model._meta.fields]
            ws.append(row_data)

        # 파일을 HttpResponse에 추가합니다.
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="selected_data_{datetime.today().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    export_selected.short_description = _("선택된 데이터 내보내기")  # 액션 버튼에 표시될 이름
    
## 실시간 날씨
@admin.register(df_weather_stts)
class MyModelAdmin_df_weather_stts(ImportExportModelAdmin,SimpleHistoryAdmin):
    list_display = ['id','file_id', 'AREA_NM','TEMP','SENSIBLE_TEMP','MAX_TEMP','MIN_TEMP','PCP_MSG','UV_MSG','AIR_MSG'] ## 보여주기
    list_filter = ['AREA_NM'] ## 필터
    search_fields = ['AREA_NM'] ## 검색
    list_per_page = 10 ##한페이지에 50개
    
    
    actions = ['export_selected']
    def export_selected(self, request, queryset):
        # queryset을 xlsx 파일로 내보냅니다.
        wb = Workbook()
        ws = wb.active

        # 필드명을 엑셀 첫 행에 추가합니다.
        field_names = [field.verbose_name.title() for field in queryset.model._meta.fields]
        ws.append(field_names)

        # 쿼리셋의 데이터를 엑셀 파일에 추가합니다.
        for obj in queryset:
            row_data = [getattr(obj, field.name) for field in queryset.model._meta.fields]
            ws.append(row_data)

        # 파일을 HttpResponse에 추가합니다.
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="selected_data_{datetime.today().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

    export_selected.short_description = _("선택된 데이터 내보내기")  # 액션 버튼에 표시될 이름